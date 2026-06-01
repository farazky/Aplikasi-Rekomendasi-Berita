import tkinter as tk                   #membuat tampilan Gui
from tkinter import ttk, font 
import os                              #mengambil lokasi file
from openpyxl import load_workbook     #membaca file excel
from datetime import datetime          #membaca waktu

#lokasi file excel
BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(BASE_DIR, "data_berita.xlsx")

#data kategori dan pewarnaanya setiap kategori
KATEGORI_LIST  = ["Politik", "Olahraga", "Teknologi", "Ekonomi", "Kesehatan"]
KATEGORI_COLOR = {
    "Politik":   "#c0392b",
    "Olahraga":  "#27ae60",
    "Teknologi": "#2980b9",
    "Ekonomi":   "#f39c12",
    "Kesehatan": "#8e44ad",
}
KATEGORI_BG = {
    "Politik":   "#fdecea",
    "Olahraga":  "#eafaf1",
    "Teknologi": "#eaf4fb",
    "Ekonomi":   "#fef9e7",
    "Kesehatan": "#f5eef8",
}

# region Excel 
#fungsi untuk membaca semua berita dari excel
def load_excel():
    wb = load_workbook(EXCEL_FILE)           #membuka file
    ws = wb["Berita"]
    berita = []
    #perulangan membaca tiap baris
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        berita.append({                     #menyimpannya menjadi dictionary     
            "id":       int(row[0]),
            "kategori": str(row[1]),
            "judul":    str(row[2]),
            "tanggal":  str(row[3])[:10] if row[3] else "",
            "views":    int(row[4]) if row[4] else 0,
        })
    wb.close()
    return berita

#menyimpan jumlah views ketika tombol baca ditekan
def save_views(berita_id, new_views):
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Berita"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == berita_id:
            row[4].value = new_views      # kolom E = Views
            break
    wb.save(EXCEL_FILE)
    wb.close()
# end region

# region Bubble Sort
# mengurutkan berita berdasarkan views terbesar
def bubble_sort_views(berita_list):
    """Sort descending by views, tie-breaker: tanggal terbaru"""
    arr = berita_list[:]
    n   = len(arr)
    for i in range(n - 1):           #loop utama
        #membandingkan item satu persatu
        for j in range(n - i - 1):
            if arr[j]["views"] < arr[j+1]["views"] or \
               (arr[j]["views"] == arr[j+1]["views"] and arr[j]["tanggal"] < arr[j+1]["tanggal"]):
                arr[j], arr[j+1] = arr[j+1], arr[j]             #pertukaran posisi
    return arr

# mengurutkan berita berdasarkan berita terbaru
def bubble_sort_tanggal(berita_list):
    """Sort descending by tanggal, tie-breaker: views terbanyak"""
    arr = berita_list[:]
    n   = len(arr)
    for i in range(n - 1):                #loop utama
        #membandingkan item satu persatu
        for j in range(n - i - 1):
            if arr[j]["tanggal"] < arr[j+1]["tanggal"] or \
               (arr[j]["tanggal"] == arr[j+1]["tanggal"] and arr[j]["views"] < arr[j+1]["views"]):
                arr[j], arr[j+1] = arr[j+1], arr[j]           #pertukaran posisi
    return arr
# end region


# region Aplikasi
class AplikasiBerita:      #Kelas utama yang mengatur seluruh tampilan aplikasi DailyNews
    def __init__(self, root):   # Mengatur konfigurasi awal aplikasi dan memanggil komponen GUI
        self.root = root
        self.root.title("DailyNews")
        self.root.geometry("1000x700")
        self.root.minsize(800, 600)
        self.root.configure(bg="#f5f0e8")
        self.aktif_kategori = KATEGORI_LIST[0]
        self.berita         = load_excel()
        self._build_fonts()
        self._build_navbar()
        self._build_pages()
        self.show_home()

    def _build_fonts(self):   # Membuat dan mengatur font yang digunakan pada aplikasi
        self.f_title   = font.Font(family="Georgia",   size=18, weight="bold")
        self.f_heading = font.Font(family="Georgia",   size=13, weight="bold")
        self.f_sub     = font.Font(family="Georgia",   size=11, weight="bold")
        self.f_small   = font.Font(family="Helvetica", size=9)
        self.f_logo    = font.Font(family="Georgia",   size=20, weight="bold")
        self.f_btn     = font.Font(family="Helvetica", size=9,  weight="bold")

    def _build_navbar(self):   # Membuat navigation bar beserta tombol navigasi
        nav = tk.Frame(self.root, bg="#1a1208", height=55)
        nav.pack(fill="x")
        nav.pack_propagate(False)
        tk.Label(nav, text="DailyNews", font=self.f_logo,
                 bg="#000000", fg="#d8b30c").pack(side="left", padx=20)
        self.btn_home = tk.Button(
            nav, text="Home", font=self.f_btn,
            bg="#1a1208", fg="#aaa", relief="flat",
            activebackground="#c0392b", activeforeground="#fff",
            cursor="hand2", padx=15, command=self.show_home)
        self.btn_home.pack(side="left")
        self.btn_kat = tk.Button(
            nav, text="Kategori", font=self.f_btn,
            bg="#1a1208", fg="#aaa", relief="flat",
            activebackground="#c0392b", activeforeground="#fff",
            cursor="hand2", padx=15, command=self.show_kategori)
        self.btn_kat.pack(side="left")

    def _set_nav_active(self, btn):   # Mengatur tombol navigasi yang sedang aktif agar tampil berbeda dari tombol lainnya
        for b in [self.btn_home, self.btn_kat]:
            b.configure(fg="#aaa", bg="#1a1208")
        btn.configure(fg="#c0392b", bg="#2a1e10")

    def _build_pages(self):   # Membuat halaman Home dan Kategori
        self.page_home     = tk.Frame(self.root, bg="#f5f0e8")
        self.page_kategori = tk.Frame(self.root, bg="#f5f0e8")

    def _hide_all(self):   # Menyembunyikan seluruh halaman yang sedang aktif
        self.page_home.pack_forget()
        self.page_kategori.pack_forget()

    def _make_scroll(self, parent):  # Membuat area scroll agar konten yang panjang tetap dapat ditampilkan
        frame  = tk.Frame(parent, bg="#f5f0e8")
        frame.pack(fill="both", expand=True)
        canvas = tk.Canvas(frame, bg="#f5f0e8", highlightthickness=0)
        sb     = ttk.Scrollbar(frame, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        inner  = tk.Frame(canvas, bg="#f5f0e8")
        wid    = canvas.create_window((0, 0), window=inner, anchor="nw")
        canvas.bind("<Configure>",      lambda e: canvas.itemconfig(wid, width=e.width))
        inner.bind("<Configure>",       lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units"))
        return inner

    def _section_title(self, parent, text, badge=""):  # Membuat judul setiap section beserta garis pemisah
        f = tk.Frame(parent, bg="#f5f0e8")
        f.pack(fill="x", pady=(15, 8))
        tk.Label(f, text=text, font=self.f_heading,
                 bg="#f5f0e8", fg="#1a1208").pack(side="left")
        if badge:
            tk.Label(f, text=f" {badge} ", font=self.f_small,
                     bg="#c0392b", fg="white", padx=5, pady=2).pack(side="left", padx=8)
        tk.Frame(parent, bg="#c0392b", height=2).pack(fill="x", pady=(0, 10))

    # region Card
    def _card(self, parent, b, rank=None, show_views=False, show_tanggal=False):   # Membuat tampilan kartu berita (card)
        color = KATEGORI_COLOR[b["kategori"]]
        bg    = KATEGORI_BG[b["kategori"]]

        card = tk.Frame(parent, bg=bg, relief="flat", bd=1,
                        cursor="hand2", width=300, height=190)
        card.pack(side="left", padx=5, expand=True, fill="x")
        card.pack_propagate(False)

        banner = tk.Frame(card, bg=color, height=80)
        banner.pack(fill="x")
        banner.pack_propagate(False)
        tk.Label(banner, text=b["emoji"], font=font.Font(size=28),
                 bg=color).pack(expand=True)
        if rank:
            tk.Label(banner, text=f"#{rank}", font=self.f_btn,
                     bg="#1a1208", fg="white", padx=4, pady=2).place(x=5, y=5)
        if show_views:
            tk.Label(banner, text=f"👁 {b['views']}x", font=self.f_small,
                     bg="#1a1208", fg="white", padx=4, pady=2).place(relx=1, x=-5, y=5, anchor="ne")

        body = tk.Frame(card, bg=bg, padx=8, pady=6)
        body.pack(fill="both", expand=True)
        tk.Label(body, text=b["kategori"].upper(), font=self.f_small,
                 bg=bg, fg=color).pack(anchor="w")
        if show_tanggal:
            tk.Label(body, text=f"📅 {b['tanggal']}", font=self.f_small,
                     bg=bg, fg="#7a6e5f").pack(anchor="w")
        judul = tk.Label(body, text=b["judul"], font=self.f_sub,
                         bg=bg, fg="#1a1208", wraplength=260, justify="left")
        judul.pack(anchor="w")
        desc = tk.Label(body, text=b["deskripsi"][:70]+"...", font=self.f_small,
                        bg=bg, fg="#7a6e5f", wraplength=260, justify="left")
        desc.pack(anchor="w")

        for w in [card, banner, body, judul, desc]:
            w.bind("<Enter>", lambda e, c=card: c.configure(relief="groove", bd=2))
            w.bind("<Leave>", lambda e, c=card: c.configure(relief="flat",   bd=1))
            w.bind("<Button-1>", lambda e, b=b: self.tambah_views(b))
    # end region

    # region List Item
    def _list_item(self, parent, b):  # Membuat tampilan berita dalam bentuk daftar pada halaman kategori
        color = KATEGORI_COLOR[b["kategori"]]
        bg    = KATEGORI_BG[b["kategori"]]

        item = tk.Frame(parent, bg=bg, relief="flat", bd=1, pady=10)
        item.pack(fill="x", pady=3)

        body = tk.Frame(item, bg=bg, padx=14)
        body.pack(side="left", fill="both", expand=True)
        tk.Label(body, text=b["judul"], font=self.f_sub,
                 bg=bg, fg="#1a1208", wraplength=700, justify="left").pack(anchor="w")

        meta = tk.Frame(body, bg=bg)
        meta.pack(anchor="w", pady=(4, 0))
        tk.Label(meta, text=f"📅 {b['tanggal']}",
                 font=self.f_small, bg=bg, fg="#aaa").pack(side="left", padx=(0, 12))
        tk.Label(meta, text=f"👁 {b['views']}x dibaca",
                 font=self.f_small, bg=bg, fg=color).pack(side="left")

        btn_frame = tk.Frame(item, bg=bg, padx=12)
        btn_frame.pack(side="right", fill="y")
        tk.Button(btn_frame, text="Baca", font=self.f_btn,
                  bg=color, fg="white", relief="flat",
                  cursor="hand2", padx=10, pady=4,
                  command=lambda b=b: self.tambah_views(b)).pack(expand=True)
    # end region

    # region Home List Item
    def _home_list_item(self, parent, b, show_rank=False, rank=None):  # Membuat tampilan daftar berita pada halaman Home
        color = KATEGORI_COLOR[b["kategori"]]
        bg    = KATEGORI_BG[b["kategori"]]

        item = tk.Frame(parent, bg=bg, relief="flat", bd=1, pady=10)
        item.pack(fill="x", pady=3)

        body = tk.Frame(item, bg=bg, padx=14)
        body.pack(side="left", fill="both", expand=True)

        # Baris atas: rank + judul
        judul_text = f"#{rank}  {b['judul']}" if show_rank else b["judul"]
        tk.Label(body, text=judul_text, font=self.f_sub,
                 bg=bg, fg="#1a1208", wraplength=700, justify="left").pack(anchor="w")

        # Baris meta: tanggal + views
        meta = tk.Frame(body, bg=bg)
        meta.pack(anchor="w", pady=(4, 0))
        tk.Label(meta, text=f"📅 {b['tanggal']}",
                 font=self.f_small, bg=bg, fg="#aaa").pack(side="left", padx=(0, 16))
        tk.Label(meta, text=f"👁 {b['views']}x dibaca",
                 font=self.f_small, bg=bg, fg=color).pack(side="left")

        # Tombol Baca (kanan)
        btn_frame = tk.Frame(item, bg=bg, padx=12)
        btn_frame.pack(side="right", fill="y")
        tk.Button(btn_frame, text="Baca", font=self.f_btn,
                  bg=color, fg="white", relief="flat",
                  cursor="hand2", padx=10, pady=4,
                  command=lambda b=b: self.tambah_views(b)).pack(expand=True)
    # end region

    #region show home
    def show_home(self):
        self._hide_all() #menyembunyikan semua halaman terlebih dahulu
        self._set_nav_active(self.btn_home) #highlight tombol home di navbar
        for w in self.page_home.winfo_children():
            w.destroy() #menghapus isi lama agar tidak menumpuk
        self.page_home.pack(fill="both", expand=True)
        inner = self._make_scroll(self.page_home) # membuat area yang bisa di scroll

        hdr = tk.Frame(inner, bg="#c0392b", pady=8)
        hdr.pack(fill="x")
        tk.Label(hdr, text="Berita Viral & Terbaru",
                 font=self.f_btn, bg="#c0392b", fg="white").pack()

        pad = tk.Frame(inner, bg="#f5f0e8")
        pad.pack(fill="x", padx=20, pady=10)

        # Jalankan bubble sort berdasarkan views, dan diambil 6 teratas
        viral = bubble_sort_views(self.berita)[:6]
        self._section_title(pad, "Berita Viral")
        for i, b in enumerate(viral): 
            self._home_list_item(pad, b, show_rank=True, rank=i+1) #tampilkan rangking di depan jdul

        # Jalankan buble sort berdasarkan tanggal, dan ambil 6 teratas
        terbaru = bubble_sort_tanggal(self.berita)[:6]
        self._section_title(pad, "Berita Terbaru")
        for b in terbaru:
            self._home_list_item(pad, b, show_rank=False)
    #end region show home
    
    #region show kategori
    def show_kategori(self):
        self._hide_all() # sembunyikan semua halaman dulu
        self._set_nav_active(self.btn_kat) # highlight tombol kategori di navbar
        for w in self.page_kategori.winfo_children():
            w.destroy() #hapus isi lama
        self.page_kategori.pack(fill="both", expand=True)

        #buat tab bar berisi tombol tiap kategori
        tab_bar = tk.Frame(self.page_kategori, bg="#1a1208", pady=5)
        tab_bar.pack(fill="x")

        self.tab_buttons = {}
        for kat in KATEGORI_LIST:
            btn = tk.Button(
                tab_bar, text=kat, font=self.f_btn,
                bg="#1a1208", fg="#aaa", relief="flat",
                activebackground=KATEGORI_COLOR[kat], activeforeground="white",
                cursor="hand2", padx=12, pady=6,
                #lambda j=kat agar nilai kat tersimpan per iterasi
                command=lambda k=kat: self._switch_kategori(k))
            btn.pack(side="left", padx=2)
            self.tab_buttons[kat] = btn

        self.kat_content = tk.Frame(self.page_kategori, bg="#f5f0e8")
        self.kat_content.pack(fill="both", expand=True)
        #tampilkan kategori yang terakhir aktif
        self._switch_kategori(self.aktif_kategori)
    #end region show kategori

    #region switch kategori
    def _switch_kategori(self, kat):
        self.aktif_kategori = kat #simpan kategori yang sedang aktif
       
       #update warna tab: aktif = warna kategori, tidak aktif = abu-abu
        for k, btn in self.tab_buttons.items():
            btn.configure(
                bg=KATEGORI_COLOR[k] if k == kat else "#1a1208",
                fg="white"           if k == kat else "#aaa")
       #hapus konten kategori sebelumnya
        for w in self.kat_content.winfo_children():
            w.destroy()

        pad = tk.Frame(self.kat_content, bg="#f5f0e8")
        pad.pack(fill="both", expand=True, padx=20, pady=15)

        color = KATEGORI_COLOR[kat]
        tk.Label(pad, text=kat, font=self.f_title,
                 bg="#f5f0e8", fg=color).pack(anchor="w")
        tk.Frame(pad, bg=color, height=3).pack(fill="x", pady=(4, 15))

        # filter berita sesuai kategori, lalu bubble sort berdasarkan views
        kat_berita = [b for b in self.berita if b["kategori"] == kat]
        sorted_kat = bubble_sort_views(kat_berita)
        for b in sorted_kat[:2]: #tampilkan 2 berita teratas
            self._list_item(pad, b)
    #end region switch kategori


    #region tambah_views
    def tambah_views(self, b):
        b["views"] += 1                 #tambah 1 views berita yang di klik
        save_views(b["id"], b["views"]) #simpan perubahan di excel
        self.berita = load_excel()      #muat ulang data dari excel
        # cek halaman yang sedang aktif, lalu refresh
        if self.page_home.winfo_ismapped():
            self.show_home()
        else:
            self.show_kategori()
    #end region tambah views

# Entry point - hanya dijalankan jika filr dieksekusi langsung 
if __name__ == "__main__":
    root = tk.Tk()             #buat jendela utama Tkinter
    AplikasiBerita(root)       #inisialisasi aplikasi
    root.mainloop()            #jalankan event loop — aplikasi menunggu interaksi user